import csv
import io
import re
from pathlib import Path


from django.db.models import Q
from rest_framework import generics, permissions, status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from .models import MatriculeEtudiantAutorise, ParametresValidation, Utilisateur
from .serializers import (
    CustomTokenSerializer,
    InscriptionEtudiantSerializer,
    MatriculeEtudiantAutoriseSerializer,
    ParametresValidationSerializer,
    PasswordChangeSerializer,
    ProfilUpdateSerializer,
    UtilisateurAdminUpdateSerializer,
    UtilisateurCreateSerializer,
    UtilisateurSerializer,
)
from apps.journal.models import JournalActivite, log

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


class IsAdminRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == Utilisateur.Role.ADMIN)


class IsAdminOrChefDept(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(
            request.user
            and request.user.is_authenticated
            and request.user.role in [Utilisateur.Role.ADMIN, Utilisateur.Role.CHEF_DEPT]
        )


class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == 200:
            user = Utilisateur.objects.get(username=request.data['username'])
            log(user, JournalActivite.TypeAction.CONNEXION, 'Connexion reussie', request.META.get('REMOTE_ADDR'))
        return response


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return Response(UtilisateurSerializer(request.user, context={'request': request}).data)

    def patch(self, request):
        serializer = ProfilUpdateSerializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UtilisateurSerializer(request.user, context={'request': request}).data)


class ProfileView(MeView):
    parser_classes = [MultiPartParser, FormParser]


class ChangePasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = PasswordChangeSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        request.user.set_password(serializer.validated_data['new_password'])
        request.user.save(update_fields=['password'])
        return Response({'detail': 'Mot de passe modifie avec succes.'})


class UtilisateurListCreateView(generics.ListCreateAPIView):
    queryset = Utilisateur.objects.all().order_by('last_name', 'first_name')
    parser_classes = [MultiPartParser, FormParser]

    def get_permissions(self):
        return [IsAdminRole()]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UtilisateurCreateSerializer
        return UtilisateurSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        role = self.request.query_params.get('role')
        search = self.request.query_params.get('search')
        actif = self.request.query_params.get('actif')
        if role:
            qs = qs.filter(role=role)
        if actif in {'true', 'false'}:
            qs = qs.filter(actif=(actif == 'true'))
        if search:
            qs = qs.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
                | Q(filiere__icontains=search)
            )
        return qs


class UtilisateurDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Utilisateur.objects.all()
    serializer_class = UtilisateurAdminUpdateSerializer
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAdminRole]

    def retrieve(self, request, *args, **kwargs):
        user = self.get_object()
        return Response(UtilisateurSerializer(user, context={'request': request}).data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        user = self.get_object()
        serializer = self.get_serializer(user, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UtilisateurSerializer(user, context={'request': request}).data)

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        user.actif = False
        user.is_active = False
        user.save(update_fields=['actif', 'is_active'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class ParametresValidationView(APIView):
    permission_classes = [IsAdminOrChefDept]

    def get(self, request):
        serializer = ParametresValidationSerializer(ParametresValidation.get_solo())
        return Response(serializer.data)

    def patch(self, request):
        if request.user.role != Utilisateur.Role.ADMIN:
            return Response({'detail': 'Seul l’administrateur peut modifier ces paramètres.'}, status=status.HTTP_403_FORBIDDEN)
        instance = ParametresValidation.get_solo()
        serializer = ParametresValidationSerializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class MatriculesAutorisesView(generics.ListAPIView):
    serializer_class = MatriculeEtudiantAutoriseSerializer
    permission_classes = [IsAdminRole]

    def get_queryset(self):
        qs = MatriculeEtudiantAutorise.objects.all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(Q(matricule__icontains=search) | Q(nom__icontains=search) | Q(prenom__icontains=search))
        return qs[:100]


class ImportMatriculesView(APIView):
    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    REQUIRED_COLUMNS = {'matricule', 'nom'}

    def normalize_header(self, value):
        value = str(value or '').strip().lower()
        replacements = {
            'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
            'à': 'a', 'â': 'a',
            'î': 'i', 'ï': 'i',
            'ô': 'o',
            'ù': 'u', 'û': 'u',
            'ç': 'c',
        }
        for source, target in replacements.items():
            value = value.replace(source, target)
        value = re.sub(r'[^a-z0-9]+', '_', value).strip('_')
        aliases = {
            'matricules': 'matricule',
            'nom_etudiant': 'nom',
            'noms': 'nom',
            'prenom_etudiant': 'prenom',
            'prenoms': 'prenom',
            'adresse_email': 'email',
            'mail': 'email',
            'tel': 'telephone',
            'telephone_1': 'telephone',
            'departement': 'departement',
            'department': 'departement',
            'specialite': 'filiere',
            'option': 'filiere',
            'filiere_option': 'filiere',
            'niveau_etude': 'niveau',
            'niveau_d_etude': 'niveau',
        }
        return aliases.get(value, value)

    def parse_niveau(self, value):
        if value in (None, ''):
            return None
        match = re.search(r'\d+', str(value))
        return int(match.group(0)) if match else None

    def clean_cell(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def rows_from_csv(self, fichier):
        text = fichier.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
        if not reader.fieldnames or len(reader.fieldnames) == 1:
            reader = csv.DictReader(io.StringIO(text), delimiter=',')
        fieldnames = [self.normalize_header(name) for name in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            normalized_row = {}
            for original_key, value in row.items():
                normalized_row[self.normalize_header(original_key)] = self.clean_cell(value)
            rows.append(normalized_row)
        return fieldnames, rows

    def rows_from_excel(self, fichier):
        if load_workbook is None:
            raise RuntimeError('Le module openpyxl n’est pas installé. Lancez : python -m pip install openpyxl')
        workbook = load_workbook(fichier, read_only=True, data_only=True)
        sheet = workbook['Import_Etudiants'] if 'Import_Etudiants' in workbook.sheetnames else workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return [], []
        fieldnames = [self.normalize_header(header) for header in headers]
        rows = []
        for values in rows_iter:
            if not values or not any(self.clean_cell(value) for value in values):
                continue
            data = {}
            for index, header in enumerate(fieldnames):
                if not header:
                    continue
                data[header] = self.clean_cell(values[index] if index < len(values) else '')
            rows.append(data)
        workbook.close()
        return fieldnames, rows

    def post(self, request):
        fichier = request.FILES.get('fichier') or request.FILES.get('csv') or request.FILES.get('excel')
        if not fichier:
            return Response({'detail': 'Ajoutez un fichier Excel (.xlsx) ou CSV.'}, status=status.HTTP_400_BAD_REQUEST)

        suffix = Path(fichier.name).suffix.lower()
        try:
            if suffix in {'.xlsx', '.xlsm'}:
                fieldnames, rows = self.rows_from_excel(fichier)
            elif suffix == '.csv':
                fieldnames, rows = self.rows_from_csv(fichier)
            else:
                return Response(
                    {'detail': 'Format non pris en charge. Utilisez un fichier .xlsx ou .csv.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except UnicodeDecodeError:
            return Response({'detail': 'CSV illisible. Enregistrez-le en CSV UTF-8.'}, status=status.HTTP_400_BAD_REQUEST)
        except RuntimeError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        normalized = {name for name in fieldnames if name}
        if not self.REQUIRED_COLUMNS.issubset(normalized):
            return Response(
                {'detail': 'Le fichier doit contenir au minimum les colonnes matricule et nom.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        skipped = 0
        for data in rows:
            matricule = data.get('matricule')
            nom = data.get('nom')
            if not matricule or not nom:
                skipped += 1
                continue
            _, was_created = MatriculeEtudiantAutorise.objects.update_or_create(
                matricule=matricule,
                defaults={
                    'nom': nom,
                    'prenom': data.get('prenom', ''),
                    'email': data.get('email', ''),
                    'departement': data.get('departement') or 'Génie Informatique',
                    'filiere': data.get('filiere') or data.get('option', ''),
                    'niveau': self.parse_niveau(data.get('niveau')),
                    'telephone': data.get('telephone', ''),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        return Response({'created': created, 'updated': updated, 'skipped': skipped, 'format': suffix.lstrip('.')})


class InscriptionEtudiantView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = InscriptionEtudiantSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(
            {
                'detail': 'Compte étudiant créé. Vous pouvez maintenant vous connecter.',
                'user': UtilisateurSerializer(user, context={'request': request}).data,
            },
            status=status.HTTP_201_CREATED,
        )
